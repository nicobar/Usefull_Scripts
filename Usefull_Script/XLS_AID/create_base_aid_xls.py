
###############################################
################# IMPORTS #####################
###############################################

import os
import pexpect
import time
import ciscoconfparse as c
# from openpyxl import load_workbook
from openpyxl.workbook import Workbook


#################################################
################# Functions #####################
#################################################


def time_string():
    ''' This function returns time in DDMMYYYY format '''

    tempo = time.gmtime()
    return str(tempo[2]) + str(tempo[1]) + str(tempo[0])


def get_command_output(node_name, cmd):
    ''' This function read devices names from file,
     connects to them and write on file output of a file '''

    cmd_telnet_bridge = 'telnet ' + BRIDGE_NAME

    cmd_telnet_node = 'telnet ' + node_name
    cmd_h = str.replace(cmd, ' ', '_')
    site = SITE[:-1]
    file_name = site + '_' + my_time + '_' + cmd_h + '.txt'

    lower_string_to_expect = node_name + '#'

    string_to_expect = str.upper(lower_string_to_expect)

    child = pexpect.spawn(cmd_telnet_bridge, encoding='utf-8')

    child.expect('login: ')
    child.sendline(MyUsername)
    child.expect('Password: ')
    child.sendline(MyBridgePwd)
    child.expect('\$')

    child.sendline(cmd_telnet_node)
    child.expect('username: ')
    child.sendline(MyUsername)
    child.expect('password: ')
    child.sendline(MyTacacsPwd)
    child.expect(string_to_expect)
    child.sendline('term len 0')
    child.expect(string_to_expect)

    child.sendline(cmd)

    with open(BASE_DIR + file_name, 'a') as fout:
        child.logfile_read = fout
        child.expect(string_to_expect)

    child.terminate()

    return file_name


def get_remote_cmd(node_name, cmd):
    ''' This function read devices names from file,
     connects to them and write on file output of a file '''

    cmd_telnet_bridge = 'telnet ' + BRIDGE_NAME

    cmd_telnet_node = 'telnet ' + node_name
    cmd_h = str.replace(cmd, ' ', '_')
    #
    file_name = node_name + '_' + cmd_h + '.txt'

    lower_string_to_expect = node_name + '#'

    string_to_expect = str.upper(lower_string_to_expect)

    child = pexpect.spawn(cmd_telnet_bridge, encoding='utf-8')

    child.expect('login: ')
    child.sendline(MyUsername)
    child.expect('Password: ')
    child.sendline(MyBridgePwd)
    child.expect('\$')

    child.sendline(cmd_telnet_node)
    child.expect('username: ')
    child.sendline(MyUsername)
    child.expect('password: ')
    child.sendline(MyTacacsPwd)
    child.expect(string_to_expect)
    child.sendline('term len 0')
    child.expect(string_to_expect)

    child.sendline(cmd)

    with open(BASE_DIR + file_name, 'w') as fout:
        child.logfile_read = fout
        child.expect(string_to_expect)

    child.terminate()

    return file_name


def from_range_to_list(range_str):
    ''' tansform '1-3' in [1,2,3] '''

    mylist = []

    h_l = range_str.split('-')
    start = int(h_l[0])
    stop = int(h_l[1])
    for x in range(start, stop + 1):
        mylist.append(x)
    return mylist


def get_indexes(text_list):

    bool_line_match = False
    for line in text_list:
        line = line.strip()
        if line == 'Port          Vlans allowed on trunk' and bool_line_match is False:
            first_index = text_list.index(line)
            bool_line_match = True
        elif line == 'Port          Vlans allowed on trunk' and bool_line_match is True:
            second_index = text_list.index(line, text_list.index(line) + 1)

    return (first_index, second_index)


def manage_OSW2OSW_allowed_list(ws, path2file):

    lst = []
    vlan_count_dict = dict()

    with open(path2file, 'r') as fin:
        text = fin.read()
    text_list = text.split('\n')
    first, second = get_indexes(text_list)

    for index in (first, second):
        # next line _ is po
        _, vlan_string = text_list[index + 1].split()
        vlan_list = vlan_string.split(',')

        for v in vlan_list:
            if v.find('-') > 0:
                help_l = from_range_to_list(v)
                for elem in help_l:
                    lst.append(int(elem))
            else:
                lst.append(int(v))

    myset = set(lst)

    for set_elem in myset:
        vlan_count_dict[set_elem] = lst.count(set_elem)

#     mycol = 1
#     max_row = len(lst) + 1
#
#     for elem, myrow in zip(lst, range(1, max_row)):
#         ws.cell(row=myrow, column=mycol, value=int(elem))

    mycol = 1
    mykeys = list(vlan_count_dict.keys())
    mykeys.sort()
    max_row = len(mykeys) + 1

    for elem, myrow in zip(mykeys, range(1, max_row)):
        ws.cell(row=myrow, column=mycol, value=int(elem))
        ws.cell(row=myrow, column=mycol + 1, value=vlan_count_dict[elem])


def manage_simple(ws, path2file):

    with open(path2file, 'r') as fin:
        myrow = 1
        for line in fin:
            line = line.strip()
            line_list = line.split()
            for elem, mycol in zip(line_list, range(1, len(line_list) + 1)):
                ws.cell(row=myrow, column=mycol, value=elem)
            myrow += 1


def manage_show_vlan_brief_onesheet(ws, path2file):

    with open(path2file, 'r') as fin:
        myrow = 1

        for line in fin:
            if len(line) > 1:
                if line[:4] != 'VLAN' and line[:4] != '----' and line[:4] != 'show':
                    line_list = line.split()
                    if line_list[0][0].isnumeric() or line_list[0] == 'show':
                        for elem, mycol in zip(line_list, range(1, len(line_list) + 1)):
                            if mycol == 1:
                                ws.cell(row=myrow, column=mycol, value=int(elem))
                            else:
                                ws.cell(row=myrow, column=mycol, value=elem)
                        myrow += 1
                    else:
                        continue
                else:
                    continue
            else:
                continue


def manage_show_vlan_brief_twosheets(wb, path2file):

    search_string = 'VLAN Name                             Status    Ports'

    with open(path2file, 'r') as fin:
        text_list = [line.strip(' \n') for line in fin.readlines()]

    first_index = text_list.index(search_string)
    second_index = text_list.index(search_string, first_index + 1)

    ws1 = wb.create_sheet(title='show_vlan_brief_OSW1', index=0)

    myrow = 1
    for line in text_list[:second_index - first_index + 1]:
        line_list = line.split()
        if len(line_list) > 0:
            if line_list[0][0].isnumeric():  # or line_list[0] == 'show':
                for elem, mycol in zip(line_list, range(1, len(line_list) + 1)):
                    if mycol == 1 and elem[0].isnumeric():
                        ws1.cell(row=myrow, column=mycol, value=int(elem))
                    else:
                        ws1.cell(row=myrow, column=mycol, value=elem)

                myrow += 1

    ws2 = wb.create_sheet(title='show_vlan_brief_OSW2', index=0)

    myrow = 1
    for line in text_list[second_index - first_index:]:
        line_list = line.split()
        if len(line_list) > 0:
            if line_list[0][0].isnumeric():  # or line_list[0] == 'show':
                for elem, mycol in zip(line_list, range(1, len(line_list) + 1)):
                    if mycol == 1 and elem[0].isnumeric():
                        ws2.cell(row=myrow, column=mycol, value=int(elem))
                    else:
                        ws2.cell(row=myrow, column=mycol, value=elem)

                myrow += 1


def get_sheet_from_filename(path):
    ''' Creates {file_name: undescored_cmd} '''

    map_file_2_sheet = {}
    for elem in os.scandir(path):
        #name = elem.name
        if elem.is_file() and elem.name[:3] != 'AID':
            result = elem.name.split('_')[2:]
            if len(result) > 1:
                cmd = '_'.join(result)[:-4]  # get rid of '.txt'
                map_file_2_sheet[elem.name] = cmd
    return map_file_2_sheet


def manage_nexus_vlan(wb, nexus_filelist2sheet):

    nexus_list = nexus_filelist2sheet.keys()
    vce_vlan_dict = dict()
    all_nexus_vlan = []

    for nexus_file in nexus_list:
        VCE_CFG_TXT_IN = BASE_DIR + nexus_file
        vce_vlan_dict[nexus_file] = get_vlan_from_cfg(VCE_CFG_TXT_IN)
        ws = wb.create_sheet(title='VLAN_ON_{}'.format(nexus_filelist2sheet[nexus_file]), index=0)
        myrow = 1
        for elem in vce_vlan_dict[nexus_file]:
            if elem:
                ws.cell(row=myrow, column=1, value=int(elem))

            myrow += 1
        all_nexus_vlan += vce_vlan_dict[nexus_file]

    all_nexus_vlan_int = [int(x) for x in all_nexus_vlan]
    ws = wb.create_sheet(title='VLAN_ON_ALL_NEXUS', index=0)
    myrow = 1
    all_nexus_vlan_set = set(all_nexus_vlan_int)
    all_nexus_vlan = list(all_nexus_vlan_set)
    all_nexus_vlan.sort()

    for elem in all_nexus_vlan:
        if elem:
            ws.cell(row=myrow, column=1, value=int(elem))
        myrow += 1


def get_vlan_from_cfg(filepath):
    ''' get vlan from VCE config '''

    vlan_list = []
    parse1 = c.CiscoConfParse(filepath)
    vlan_rough = parse1.find_lines(r'^vlan')

    for v in vlan_rough:
        vlan_list.append(v.split()[1])

    return vlan_list


def manage_static_routes(wb, nexus_file_list):

    text_list = []
    ws = wb.create_sheet(title='VCE_Static_Routes', index=0)
    for file in nexus_file_list:
        if 'VSW' not in file:
            parse1 = c.CiscoConfParse(BASE_DIR + file)
            routes_rough = parse1.find_lines(r'^ ip route')
            routes_rough_h = [x for x in routes_rough if 'ip router ospf' not in x]
            text_list += routes_rough_h
    myrow = 1

    for line in text_list:
        line_list = line.split()
        for elem, mycol in zip(line_list, range(1, len(line_list) + 1)):
            ws.cell(row=myrow, column=mycol, value=elem)
        myrow += 1


def manage_rb(wb, node_list):

    mac_osw_map = dict()  # {mac: name}
    vlan_rb_map = dict()  # {vlan: mac}
    new_vlan_rb_map = dict()
    ws = wb.create_sheet(title='Root-bridge per VLAN', index=0)

    for node in node_list:
        mac_osw_map[get_switch_mac_address(node)] = node
        vlan_rb_map.update(get_rb_per_vlan(node))

    for vlan in vlan_rb_map.keys():
        if vlan_rb_map[vlan] in mac_osw_map.keys():
            new_vlan_rb_map[vlan] = mac_osw_map[vlan_rb_map[vlan]]
        else:
            new_vlan_rb_map[vlan] = vlan_rb_map[vlan]

    myrow = 1
    for vlan in new_vlan_rb_map.keys():
        ws.cell(row=myrow, column=1, value=int(vlan))
        ws.cell(row=myrow, column=2, value=new_vlan_rb_map[vlan])
        myrow += 1


def get_switch_mac_address(osw):
    ''' return a string containing mac address of osw '''

    cmd = 'show spanning-tree bridge address'

    file_name = get_remote_cmd(osw, cmd)
    lst = from_file_to_cfg_as_list(file_name)
    if lst is not None:
        mac = lst[1].split()[1]
    else:
        mac = None
    return mac


def from_file_to_cfg_as_list(file_name):
    ''' return a list containing text of file_name '''
    show_cmd = []

    for elem in open(BASE_DIR + file_name, 'r'):
        show_cmd.append(elem.rstrip())
    return show_cmd


def get_rb_per_vlan(osw):
    ''' return a map {vlan: mac} indicating RB for osw '''

    cmd = 'show spanning-tree root brief'

    file_name = get_remote_cmd(osw, cmd)
    show_list = from_file_to_cfg_as_list(file_name)

    mp = {}

    for elem in show_list:
        if len(elem) > 0:
            if elem[:2] == 'VL':
                lst_elem = elem.split()
                vlan = lst_elem[0]
                mac = lst_elem[2]
                mp[vlan[4:].lstrip('0')] = mac
            else:
                continue
        else:
            continue
    return mp


def get_dot1q_from_dev_cfg(dev_file, intf):
    ''' Returns a list of dot1q tag taken on VPE from intf'''

    parse = c.CiscoConfParse(dev_file)
    be_obj = parse_osw.find_objects(r'^interface ' + intf)
    be_cfg = be_obj[0].ioscfg

    for line in be_cfg:
        s = line.split(' ')
        if len(s) >= 5:
            if s[3] == 'allowed' and s[5][0].isdigit():
                help_str += s[5] + ','
            elif s[3] == 'allowed' and s[5] == 'add':
                help_str += s[6] + ','

    help_list = help_str[:-1].split(',')

    for elem in help_list:
        if elem.find('-') > 0:
            help_l = from_range_to_list(elem)
            for elemh in help_l:
                osw_vlan_set.add(elemh)
        else:
            osw_vlan_set.add(elem)

    result_l = list(osw_vlan_set)

    result_l.sort(key=natural_keys)
    result_s = ','.join(result_l)

    return result_s


def manage_dot1q(wb, path, vpeaddendum_cfg_list, vceaddendum_cfg_list):
    ''' Returns a list of dot1q tag taken on VCE om PO'''
    vpeadd_2_vpe_map = {vpeaddendum_cfg_list[0]: 'VPE1', vpeaddendum_cfg_list[1]: 'VPE2'}  # {vpeadd: 'vpe1'}
    vceadd_2_vpe_map = {vceaddendum_cfg_list[0]: 'VCE1', vceaddendum_cfg_list[1]: 'VCE2'}  # {vceadd: 'vce1'}

    vpeadd_dot1q_map = {}  # {vceadd: [list_dot1q]}
    vceadd_dot1q_map = {}  # {vpeadd: [list_dot1q]}

    for vpe in vpeaddendum_cfg_list:
        vpe_dot1q_map[vpeadd_2_vpe_map[vpe]] = get_dot1q_from_dev_cfg(path + vpe, be)  # {'VPEx': [list_dot1q]}

    for vce in vceaddendum_cfg_list:
        vce_dot1q_map[vceadd_2_vpe_map[vce]] = get_dot1q_from_dev_cfg(path + vce, po)   # {'VCEx': [list_dot1q]}

    for vpe in vpe_dot1q_map.keys():
        ws = wb.create_sheet(title='Dot1q Tag on {}'.format(vpe), index=0)
        myrow = 1
        for tag in vpe_dot1q_map[vpe]:
            ws.cell(row=myrow, column=1, value=int(tag))
            myrow += 1

    for vce in vce_dot1q_map.keys():
        ws = wb.create_sheet(title='Dot1q Tag on {}'.format(vce), index=0)
        myrow = 1
        for tag in vce_dot1q_map[vce]:
            ws.cell(row=myrow, column=1, value=int(tag))
            myrow += 1


def create_xlsx(path, site):
    OUTPUT_XLS = path + 'AID_to_{}_NMP.xlsx'.format(site)
    wb = Workbook()
    map_file_2_sheet = get_sheet_from_filename(path)

    for file, sheet in zip(map_file_2_sheet.keys(), map_file_2_sheet.values()):
        ws = wb.create_sheet(title=sheet, index=0)
        cmd_list = sheet.split('_')
        if 'trunk' in cmd_list:
            manage_OSW2OSW_allowed_list(ws, path + file)
        elif 'vlan' in cmd_list and 'brief' in cmd_list:
            manage_show_vlan_brief_onesheet(ws, path + file)
            manage_show_vlan_brief_twosheets(wb, path + file)
        elif 'standby' in cmd_list or 'vrrp' in cmd_list or 'description' in cmd_list:
            manage_simple(ws, path + file)

    manage_nexus_vlan(wb, nexus_filelist2sheet)
    manage_static_routes(wb, nexus_file_list)
    manage_rb(wb, node_list)
    manage_dot1q(wb, path, vpeaddendum_cfg_file, nexus_file_list)
    wb.save(filename=OUTPUT_XLS)


#############################################
################# VARIABLES #################
#############################################


PO = 'po1'
BASE = '/mnt/hgfs/VM_shared/VF-2017/NMP/'
SITE = 'BO01/'
BASE_DIR = BASE + SITE + 'AID/'

BRIDGE_NAME = '10.192.10.8'
MyUsername = 'zzasp70'
MyBridgePwd = "SPra0094"
MyTacacsPwd = "0094SPra_"
command_list = ['show interfaces description',
                'show vlan brief',
                'show standby brief',
                'show vrrp brief',
                'show interfaces {} trunk'.format(PO)]
#'show spanning-tree bridge address',
#'show spanning-tree root brief']

be = 'Bundle-Ether411'
po = 'Port-Channel411'

node_list = ['BOOSW013',
             'BOOSW016']

vpeaddendum_cfg_file = [node_list[0] + 'VPE_addendum.txt',
                        node_list[1] + 'VPE_addendum.txt']


nexus_filelist2sheet = {node_list[0] + 'VCE.txt': 'VCE1',
                        node_list[1] + 'VCE.txt': 'VCE2',
                        node_list[0] + 'VSW.txt': 'VSW'
                        }


nexus_file_list = nexus_filelist2sheet.keys()

############################################
################# MAIN #####################
############################################


my_time = time_string()

file_list = []
for cmd in command_list:
    for node in node_list:
        file_list.append(get_command_output(node, cmd))
# file_list = ['BO01_20112017_show_interfaces_description.txt',
#              'BO01_20112017_show_interfaces_po1_trunk.txt',
#              'BO01_20112017_show_standby_brief.txt',
#              'BO01_20112017_show_vlan_brief.txt',
#              'BO01_20112017_show_vrrp_brief.txt']

create_xlsx(BASE_DIR, SITE[:-1])
