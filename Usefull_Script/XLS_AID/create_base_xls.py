
###############################################
################# IMPORTS #####################
###############################################

import os
import pexpect
import time
# from openpyxl import load_workbook
from openpyxl.workbook import Workbook


#################################################
################# Functions #####################
#################################################


def time_string():
    ''' This function returns time in DDMMYYYY format '''

    tempo = time.gmtime()
    return str(tempo[2]) + str(tempo[1]) + str(tempo[0])


def get_command_output(node_name, cmd):
    ''' This function read devices names from file,
     connects to them and write on file output of a file '''

    cmd_telnet_bridge = 'telnet ' + BRIDGE_NAME

    cmd_telnet_node = 'telnet ' + node_name
    cmd_h = str.replace(cmd, ' ', '_')
    site = SITE[:-1]
    file_name = site + '_' + my_time + '_' + cmd_h + '.txt'

    lower_string_to_expect = node_name + '#'

    string_to_expect = str.upper(lower_string_to_expect)

    child = pexpect.spawn(cmd_telnet_bridge, encoding='utf-8')

    child.expect('login: ')
    child.sendline(MyUsername)
    child.expect('Password: ')
    child.sendline(MyBridgePwd)
    child.expect('\$')

    child.sendline(cmd_telnet_node)
    child.expect('username: ')
    child.sendline(MyUsername)
    child.expect('password: ')
    child.sendline(MyTacacsPwd)
    child.expect(string_to_expect)
    child.sendline('term len 0')
    child.expect(string_to_expect)

    child.sendline(cmd)

    with open(BASE_DIR + file_name, 'a') as fout:
        child.logfile_read = fout
        child.expect(string_to_expect)

    child.terminate()

    return file_name


def from_range_to_list(range_str):
    ''' tansform '1-3' in [1,2,3] '''

    mylist = []

    h_l = range_str.split('-')
    start = int(h_l[0])
    stop = int(h_l[1])
    for x in range(start, stop + 1):
        mylist.append(x)
    return mylist


def get_indexes(text_list):

    bool_line_match = False
    for line in text_list:
        line = line.strip()
        if line == 'Port          Vlans allowed on trunk' and bool_line_match is False:
            first_index = text_list.index(line)
            bool_line_match = True
        elif line == 'Port          Vlans allowed on trunk' and bool_line_match is True:
            second_index = text_list.index(line, text_list.index(line) + 1)

    return (first_index, second_index)


def manage_OSW2OSW_allowed_list(ws, path2file):

    lst = []

    with open(path2file, 'r') as fin:
        text = fin.read()
    text_list = text.split('\n')
    first, second = get_indexes(text_list)

    for index in (first, second):
        # next line _ is po
        _, vlan_string = text_list[index + 1].split()
        vlan_list = vlan_string.split(',')

        for v in vlan_list:
            if v.find('-') > 0:
                help_l = from_range_to_list(v)
                for elem in help_l:
                    lst.append(int(elem))
            else:
                lst.append(int(v))

    mycol = 1
    max_row = len(lst) + 1

    for elem, myrow in zip(lst, range(1, max_row)):
        ws.cell(row=myrow, column=mycol, value=int(elem))


def manage_simple(ws, path2file):

    with open(path2file, 'r') as fin:
        myrow = 1
        for line in fin:
            line = line.strip()
            line_list = line.split()
            for elem, mycol in zip(line_list, range(1, len(line_list) + 1)):
                ws.cell(row=myrow, column=mycol, value=elem)
            myrow += 1


def manage_show_vlan_brief(ws, path2file):

    with open(path2file, 'r') as fin:
        myrow = 1

        for line in fin:
            if len(line) > 1:
                if line[:4] != 'VLAN' and line[:4] != '----' and line[:4] != 'show':
                    line_list = line.split()
                    if line_list[0][0].isnumeric() or line_list[0] == 'show':
                        for elem, mycol in zip(line_list, range(1, len(line_list) + 1)):
                            ws.cell(row=myrow, column=mycol, value=elem)
                        myrow += 1
                    else:
                        continue
                else:
                    continue
            else:
                continue


def get_sheet_from_filename(path):
    ''' Creates {file_name: undescored_cmd} '''

    map_file_2_sheet = {}
    for elem in os.scandir(path):
        #name = elem.name
        if elem.is_file() and elem.name[:3] != 'AID':
            result = elem.name.split('_')[2:]
            cmd = '_'.join(result)[:-4]  # get rid of '.txt'
            map_file_2_sheet[elem.name] = cmd
    return map_file_2_sheet


def create_xlsx(path, site):

    OUTPUT_XLS = path + 'AID_to_{}_NMP.xlsx'.format(site)
    wb = Workbook()
    map_file_2_sheet = get_sheet_from_filename(path)

    for file, sheet in zip(map_file_2_sheet.keys(), map_file_2_sheet.values()):
        ws = wb.create_sheet(title=sheet, index=0)
        cmd_list = sheet.split('_')
        if 'trunk' in cmd_list:
            manage_OSW2OSW_allowed_list(ws, path + file)
        elif 'vlan' in cmd_list and 'brief' in cmd_list:
            manage_show_vlan_brief(ws, path + file)
        else:
            manage_simple(ws, path + file)

    wb.save(filename=OUTPUT_XLS)

#############################################
################# VARIABLES #################
#############################################


PO = 'po1'
BASE = '/mnt/hgfs/VM_shared/VF-2017/NMP/'
SITE = 'BO01/'
BASE_DIR = BASE + SITE + 'AID_2/'
BRIDGE_NAME = '10.192.10.8'
MyUsername = 'zzasp70'
MyBridgePwd = "SP9400ra"
MyTacacsPwd = "0094SPra_"
command_list = ['show interfaces description',
                'show vlan brief',
                'show standby brief',
                'show vrrp brief',
                'show interfaces {} trunk'.format(PO)]

node_list = ['BOOSW013',
             'BOOSW016']


############################################
################# MAIN #####################
############################################


my_time = time_string()

file_list = []
for cmd in command_list:
    for node in node_list:
        file_list.append(get_command_output(node, cmd))


create_xlsx(BASE_DIR, SITE[:-1])
